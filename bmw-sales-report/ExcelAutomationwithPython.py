
from pathlib import Path
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

# CONFIG
CSV_PATH   = "BMW sales data (2010-2024) (1).csv"
OUT_XLSX   = "bmw_sales_report.xlsx"
OUT_DIR    = Path(".")  # charts will be saved here
STYLE      = "ggplot"   # nicer defaults (built-in matplotlib style)
DPI        = 180        # high resolution for portfolio
plt.style.use(STYLE)

# HELPERS
def kfmt(x, _):
    """Thousands formatter, e.g. 12,300,000 -> 12.3M"""
    absx = abs(x)
    if absx >= 1e9:  return f"{x/1e9:.1f}B"
    if absx >= 1e6:  return f"{x/1e6:.1f}M"
    if absx >= 1e3:  return f"{x/1e3:.1f}K"
    return f"{x:.0f}"

def bar_label(ax):
    """Attach value labels to bars (works for bar and barh)."""
    try:
        for c in ax.containers:
            ax.bar_label(c, fontsize=9, padding=3, fmt="%.0f")
    except Exception:
        pass

def save_fig(name):
    path = OUT_DIR / name
    plt.tight_layout()
    plt.savefig(path, dpi=DPI, bbox_inches="tight")
    plt.close()
    return path

# LOAD & CLEAN
df = pd.read_csv(CSV_PATH)

expected = ["Year", "Region", "Model", "Sales_Volume", "Price_USD", "Fuel_Type"]
missing  = [c for c in expected if c not in df.columns]
if missing:
    print("⚠️ Missing columns:", missing)

for col in ["Year", "Sales_Volume", "Price_USD"]:
    if col in df.columns:
        df[col] = pd.to_numeric(df[col], errors="coerce")

# PIVOTS
sheets = {}

if {"Year","Sales_Volume"} <= set(df.columns):
    yearly = (df.dropna(subset=["Year"])
                .groupby("Year", as_index=False)["Sales_Volume"].sum()
                .sort_values("Year"))
    sheets["Yearly_Sales"] = yearly

if {"Region","Sales_Volume"} <= set(df.columns):
    region = (df.groupby("Region", as_index=False)["Sales_Volume"].sum()
                .sort_values("Sales_Volume", ascending=False))
    sheets["Sales_by_Region"] = region

if {"Model","Sales_Volume"} <= set(df.columns):
    top_models = (df.groupby("Model", as_index=False)["Sales_Volume"].sum()
                    .sort_values("Sales_Volume", ascending=False)
                    .head(10))
    sheets["Top_10_Models"] = top_models

if {"Fuel_Type","Price_USD"} <= set(df.columns):
    fuel = (df.groupby("Fuel_Type", as_index=False)["Price_USD"].mean()
              .round(2).sort_values("Price_USD", ascending=False))
    sheets["Avg_Price_by_Fuel"] = fuel

# CHARTS
charts = {}

if "Yearly_Sales" in sheets:
    s = sheets["Yearly_Sales"]
    fig, ax = plt.subplots(figsize=(8, 4.5))
    ax.plot(s["Year"], s["Sales_Volume"], marker="o", linewidth=2)
    ax.set_title("Yearly Sales Volume", fontsize=13)
    ax.set_xlabel("Year"); ax.set_ylabel("Sales Volume")
    ax.yaxis.set_major_formatter(FuncFormatter(kfmt))
    ax.grid(True, alpha=.3)
    charts["yearly_sales.png"] = save_fig("yearly_sales.png")

if "Sales_by_Region" in sheets:
    s = sheets["Sales_by_Region"]
    fig, ax = plt.subplots(figsize=(8, 4.8))
    ax.barh(s["Region"], s["Sales_Volume"])
    ax.set_title("Sales by Region", fontsize=13)
    ax.set_xlabel("Sales Volume")
    ax.xaxis.set_major_formatter(FuncFormatter(kfmt))
    ax.invert_yaxis()
    bar_label(ax)
    charts["sales_by_region.png"] = save_fig("sales_by_region.png")

if "Top_10_Models" in sheets:
    s = sheets["Top_10_Models"].sort_values("Sales_Volume")  # for nicer barh order
    fig, ax = plt.subplots(figsize=(8, 5))
    ax.barh(s["Model"], s["Sales_Volume"])
    ax.set_title("Top 10 Models by Sales", fontsize=13)
    ax.set_xlabel("Sales Volume")
    ax.xaxis.set_major_formatter(FuncFormatter(kfmt))
    ax.invert_yaxis()
    bar_label(ax)
    charts["top_10_models.png"] = save_fig("top_10_models.png")

if "Avg_Price_by_Fuel" in sheets:
    s = sheets["Avg_Price_by_Fuel"]
    fig, ax = plt.subplots(figsize=(7.2, 4))
    ax.bar(s["Fuel_Type"], s["Price_USD"])
    ax.set_title("Average Price by Fuel Type (USD)", fontsize=13)
    ax.set_xlabel("Fuel Type"); ax.set_ylabel("Average Price (USD)")
    ax.yaxis.set_major_formatter(FuncFormatter(lambda v, _: f"${v:,.0f}"))
    bar_label(ax)
    plt.xticks(rotation=10)
    charts["avg_price_by_fuel.png"] = save_fig("avg_price_by_fuel.png")

# WRITE EXCEL 
with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Data", index=False)
    for name, sdf in sheets.items():
        sdf.to_excel(writer, sheet_name=name, index=False)

# Embed charts into a Dashboard sheet
wb = load_workbook(OUT_XLSX)
if "Dashboard" in wb.sheetnames:
    ws = wb["Dashboard"]
else:
    ws = wb.create_sheet("Dashboard", 0)

ws["A1"] = "BMW Sales Dashboard"
ws["A1"].font = ws["A1"].font.copy(bold=True)

# Place images in a grid (requires Pillow installed)
row_anchor = {"yearly_sales.png":"A3",
              "sales_by_region.png":"J3",
              "top_10_models.png":"A22",
              "avg_price_by_fuel.png":"J22"}

for fname, anchor in row_anchor.items():
    if fname in charts:
        img = XLImage(str(charts[fname]))
        ws.add_image(img, anchor)

wb.save(OUT_XLSX)

print(f"Report created: {OUT_XLSX}")
print("Charts saved:", ", ".join(charts.keys()))
print("Dashboard sheet contains embedded charts.")